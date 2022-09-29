import openpyxl

#Criar a planilha
book = openpyxl.Workbook()

#Como vizualizar páginas existentes
print(book.sheetnames)

#Como criar a página
book.create_sheet('Frutas')

#Selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])

#Adicionar dados à página
frutas_page.append(['Banana', '5', 'R$ 3,90'])
frutas_page.append(['Maçã', '2', 'R$ 4,50'])
frutas_page.append(['Goiaba', '10', 'R$ 1,90'])
frutas_page.append(['Laranja', '2', 'R$ 2,60']) 

#Salvar a planilha
book.save('Planilha de compras.xlsx')

